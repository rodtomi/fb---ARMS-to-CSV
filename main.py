# TO CREATE A NEW .EXE TYPE THE FOLLOWING COMMAND IN THE TERMINAL:
#     pyinstaller main.py -F --onefile
# The .exe file will be found inside the dist directory.
# (you need to have pyinstaller previously installed)
#
# OR you can use auto-py-to-exe by executing, in the terminal:
# auto-py-to-exe
#

import csv
import datetime
import sys
from datetime import datetime
from tkinter import *
from tkinter import filedialog

import openpyxl

# Get current file's directory:
#dir_path = os.path.dirname(os.path.realpath(__file__))
dir_path = sys.executable
print(dir_path)

def btn_Browse_clicked(t_entry):
    filename = filedialog.askopenfilename(initialdir=dir_path, title="Seleccionar archivo ...")
    if filename != "":
        t_entry.delete(0,"end")
        t_entry.insert(0,filename)


def btn_Check_clicked():
    # Declaring constants for source workbook ...
    col_DateOfDuty = 2
    col_FlightDetails = 4
    col_ReportingDateTime = 5
    col_StartTime = 6
    col_EndTime = 7
    col_TrgRemarks = 8

    # Initializing variables to be used in csv file ...
    r_Subject = 'Subject'
    r_Start_Date = 'Start Date'
    r_Start_Time = 'Start Time'
    r_End_Date = 'End Date'
    r_End_Time = 'End Time'
    r_All_Day_Event = 'All Day Event'
    r_Description = 'Description'
    r_Location = 'Location'
    r_Private = 'Private'
    r_Row = [r_Subject, r_Start_Date, r_Start_Time, r_End_Date, r_End_Time, r_All_Day_Event, r_Description, r_Location,
             r_Private]

    # Opening Workbooks & Files ...
    # wb_calendar = Workbook()
    source_file_and_path = entry_Path.get()
    # wb_source = openpyxl.load_workbook("Roster.xlsx")
    wb_source = openpyxl.load_workbook(source_file_and_path)

    csv_file_name = 'googleCalendar-' + str(datetime.now().strftime("%Y%m%d%H%M%S")) + '.csv'
    csv_file = open(csv_file_name, 'w', encoding='UTF8', newline='')
    # creating a csv writer object
    csv_file_writer = csv.writer(csv_file)
    # write the first line (headers) ...
    csv_file_writer.writerow(r_Row)

    # Selecting the sheet = sh from the source file from ARMS...
    # sh = wb_source['Roster']
    sh = wb_source.worksheets[0]
    max_rows = sh.max_row

    # Positioning on the first Cell for 'Reporting Date Time' ...
    curr_row = 3
    curr_cell = sh.cell(curr_row, col_ReportingDateTime)
    # print(curr_cell.value)

    while max_rows > curr_row:
        curr_row += 1
        curr_cell = sh.cell(curr_row, col_ReportingDateTime)
        v_Description = ''
        # If cell value is blank, go to the next row ...
        while curr_cell.value == None:
            curr_row += 1
            if curr_row > max_rows:
                break
            curr_cell = sh.cell(curr_row, col_ReportingDateTime)
            # If exceeded sheet range, quit loop
        if curr_row > max_rows:
            break

        # Once encountered a cell with a value, move to the next one which has date/time format in local (L) time ...
        curr_row += 1
        curr_cell = sh.cell(curr_row, col_ReportingDateTime)  # REPORTING DATE TIME

        # ==========================================================================
        v_Subject = str(sh.cell(curr_row - 1, col_FlightDetails).value) + ' ' + str(
            sh.cell(curr_row, col_FlightDetails).value)
        if v_Subject == "None None":
            v_Subject = str(sh.cell(curr_row - 1, col_TrgRemarks).value.replace(",", " - "))
        v_Description = v_Subject
        # ==========================================================================

        # print(curr_cell.value)

        # Looking for the row with the cell with the latest End Time ...
        aux_delimCounter = 2
        temp_row = curr_row
        while (sh.cell(temp_row + 1, col_ReportingDateTime).value == None) and (
                sh.cell(temp_row + 1, col_EndTime).value != None) and (
                sh.cell(temp_row + 1, col_FlightDetails).value != None):
            temp_row += 1
            if temp_row > max_rows:
                break
            if (aux_delimCounter == 2):
                v_Description += ' / '
                aux_delimCounter = 0
            v_Description += ' ' + str(sh.cell(temp_row, col_FlightDetails).value)
            aux_delimCounter += 1
            # Once encountered, we move to the previous one which has date/time format in local (L) time ...
            # temp_row -= 1

        # print(sh.cell(temp_row, col_EndTime).value)

        # ---------------------------------------------------------------------------------------------------------
        # Populate new Calendar workbook with all the data ...
        # ---------------------------------------------------------------------------------------------------------
        # SUBJECT:
        # dest_sh.cell(dest_curr_row, dest_col_Subject).value = v_Subject
        r_Subject = v_Subject
        # START DATE:
        last_DateDuty = datetime.strptime(str(sh.cell(4, col_DateOfDuty).value), '%Y-%m-%d %H:%M:%S')
        current_DateDuty_temp = sh.cell(curr_row - 1, col_DateOfDuty).value
        if current_DateDuty_temp == None:
            current_DateDuty = last_DateDuty
        else:
            current_DateDuty = datetime.strptime(str(sh.cell(curr_row - 1, col_DateOfDuty).value), '%Y-%m-%d %H:%M:%S')
            last_DateDuty = current_DateDuty
        current_year = current_DateDuty.strftime("%Y")
        v_StartDateTime = datetime.strptime(
            sh.cell(curr_row, col_ReportingDateTime).value.rstrip('(L)') + " " + str(current_year), '%d-%b %H:%M %Y')
        # dest_sh.cell(dest_curr_row, dest_col_StartDate).value = v_StartDateTime.strftime("%m/%d/%Y")
        r_Start_Date = v_StartDateTime.strftime("%m/%d/%Y")
        # START TIME:
        # dest_sh.cell(dest_curr_row, dest_col_StartTime).value = v_StartDateTime.strftime("%H:%M")
        r_Start_Time = v_StartDateTime.strftime("%H:%M")
        # END DATE:
        if (v_StartDateTime.strftime("%m") == 12) and (
                datetime.strptime(sh.cell(temp_row, col_EndTime).value.rstrip('(L)'), '%d-%b %H:%M').strftime(
                        "%m") == 1):  # si llega al año siguiente ...
            current_year += 1
        v_EndDateTime = datetime.strptime(sh.cell(temp_row, col_EndTime).value.rstrip('(L)') + " " + str(current_year),
                                          '%d-%b %H:%M %Y')
        # dest_sh.cell(dest_curr_row, dest_col_EndDate).value = v_EndDateTime.strftime("%m/%d/%Y")
        r_End_Date = v_EndDateTime.strftime("%m/%d/%Y")
        # END TIME:
        # dest_sh.cell(dest_curr_row, dest_col_EndTime).value = v_EndDateTime.strftime("%H:%M")
        r_End_Time = v_EndDateTime.strftime("%H:%M")
        # ALL DAY EVENT:
        # dest_sh.cell(dest_curr_row, dest_col_AllDayEvent).value = 'False'
        r_All_Day_Event = 'False'
        # DESCRIPTION:
        # dest_sh.cell(dest_curr_row, dest_col_Description).value = v_Description
        r_Description = v_Description
        # LOCATION:
        v_ReportingLocation = v_Subject[7:10]
        if (v_ReportingLocation == 'AEP'):
            # dest_sh.cell(dest_curr_row, dest_col_Location).value = 'Aeroparque Internacional Jorge Newbery'
            r_Location = 'Aeroparque Internacional Jorge Newbery'
        elif (v_ReportingLocation == 'EZE'):
            # dest_sh.cell(dest_curr_row, dest_col_Location).value = 'Ezeiza International Airport'
            r_Location = 'Ezeiza International Airport'
        else:
            # dest_sh.cell(dest_curr_row, dest_col_Location).value = ''
            r_Location = ''
        # PRIVATE:
        # dest_sh.cell(dest_curr_row, dest_col_Private).value = 'False'
        r_Private = 'False'

        # Write row in csv
        r_Row = [r_Subject, r_Start_Date, r_Start_Time, r_End_Date, r_End_Time, r_All_Day_Event, r_Description,
                 r_Location,
                 r_Private]
        csv_file_writer.writerow(r_Row)

        # Go to next Row in destination file ...
        # dest_curr_row += 1
        # ---------------------------------------------------------------------------------------------------------
    csv_file.close()
    del csv_file_writer

# INICIALIZO VENTANA

window = Tk()
window.iconbitmap('paperplane_116175.ico')
window.title("flybondi: ARMS to .CSV")
window.geometry("1000x600")
window.configure(bg = "#f6f6f6")
canvas = Canvas(
    window,
    bg = "#f6f6f6",
    height = 600,
    width = 1000,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge")
canvas.place(x = 0, y = 0)

# INICIALIZO FONDO
background_img = PhotoImage(file = f"background.png")
background = canvas.create_image(
    500.0, 300.0,
    image=background_img)

# INICIALIZO TEXT BOX (TIPO 'ENTRY') DE FILE PATH:
entryPath_img = PhotoImage(file = f"img_textBox0.png")
entryPath_bg = canvas.create_image(
    672.0, 188.5,
    image = entryPath_img)
entry_Path = Entry(
    bd = 0,
    bg = "#ffffff",
    highlightthickness = 0)
entry_Path.place(
    x = 466.5, y = 171,
    width = 411.0,
    height = 37)
entry_Path.insert(0, "Seleccionar archivo ...")

# INICIALIZO BOTON BROWSE:
program_loop_counter = 0
img_btnBrowse = PhotoImage(file = f"img1.png")
button_Browse = Button(
    image = img_btnBrowse,
    borderwidth = 0,
    highlightthickness = 0,
    command = lambda: btn_Browse_clicked(entry_Path),
    relief = "flat")
button_Browse.place(
    x = 904, y = 171,
    width = 39,
    height = 39)

# INICIALIZO BOTON CHECK!
img_btnCheck = PhotoImage(file = f"img0.png")
button_Check = Button(
    image = img_btnCheck,
    borderwidth = 0,
    highlightthickness = 0,
    command = lambda: btn_Check_clicked(),
    relief = "flat")
button_Check.place(
    x = 639, y = 373,
    width = 103,
    height = 42)

window.resizable(False, False)
window.mainloop()
